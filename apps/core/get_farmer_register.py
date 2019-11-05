from openpyxl import load_workbook
from organic_farmer.models import OrganicFarmerAddress, OrganicFarmerContact, EntityType, Entity, Scope, Activity, OrganicFarmer
from nhanduti.settings import STATIC_ROOT
from geopy.geocoders import Nominatim
from django.contrib.gis.geos import Point


def get_farmer_register_file():
    farmer_register_file = STATIC_ROOT + '/upload/CNPO_MAPA_31_07_2019.xlsx'
    return(farmer_register_file)


def get_farmer_register_sheet(farmer_register_file):
    farmer_register_ws = load_workbook(farmer_register_file)
    farmer_register_sheet = farmer_register_ws.get_active_sheet()
    return(farmer_register_sheet)


def get_farmer_register_titles(farmer_register_sheet):
    register_fields = {}
    for row in farmer_register_sheet.iter_rows(min_row=3, max_col=10, max_row=3, values_only=True):
        titles = row
    for title in titles:
        register_fields.update({title: ''})
    return(register_fields)


def set_activity(farmer_register_activities):
    inserir_atividades_farmer = []
    buscar_atividade = []
    criar_cadastro_atividade = []
    for atividade in farmer_register_activities:
        atividade_in_db = Activity.objects.filter(name__icontains=atividade.strip().upper())
        if atividade_in_db:
            inserir_atividades_farmer.append(atividade_in_db)
        else:
            buscar_atividade.append(atividade.strip().upper())
            criar_cadastro_atividade.append(Activity(name=atividade.strip().upper()))
    if criar_cadastro_atividade:
        Activity.objects.bulk_create(criar_cadastro_atividade)
    if buscar_atividade:
        inserir_atividades_farmer = set_activity(buscar_atividade)
        return(inserir_atividades_farmer)
    else:
        return(inserir_atividades_farmer)


def set_scope(farmer_register_scopes):
    inserir_scope_farmer = []
    buscar_scope = []
    criar_cadastro_scope = []
    for escopo in farmer_register_scopes:
        if escopo.strip().upper() is ('POV' or 'P.O.V' or 'processador de alimentos'):
            escopo_ = 'PROCESSAMENTO DE PRODUTOS DE ORIGEM VEGETAL'
        elif escopo.strip().upper() is ('POA' or 'P.O.A'):
            escopo_ = 'PROCESSAMENTO DE PRODUTOS DE ORIGEM ANIMAL'
        elif escopo.strip().upper() is ('PPV' or 'P.P.V' or 'produção vegetal'):
            escopo_ = 'PRODUÇÃO PRIMÁRIA VEGETAL'
        elif escopo.strip().upper() is ('PPA' or 'P.P.A'):
            escopo_ = 'PRODUÇÃO PRIMÁRIA ANIMAL'
        elif escopo.strip().upper() is ('ESO' or 'E.S.O' or 'EXT'):
            escopo_ = 'EXTRATIVISMO SUSTENTÁVEL ORGÂNICO'
        elif escopo.strip().upper() is ('PIA' or 'P.I.A'):
            escopo_ = 'PROCESSAMENTO DE INSUMOS AGRÍCOLAS'

        scope_in_db = Scope.objects.filter(name__icontains=escopo_.strip().upper())
        if scope_in_db:
            inserir_scope_farmer.append(scope_in_db)
        else:
            buscar_scope.append(escopo.strip().upper())
            criar_cadastro_scope.append(Activity(name=escopo.strip().upper()))
    if criar_cadastro_scope:
        Activity.objects.bulk_create(criar_cadastro_scope)
    if buscar_scope:
        inserir_scope_farmer = set_activity(buscar_scope)
        return(inserir_scope_farmer)
    else:
        return(inserir_scope_farmer)


def create_farmer_register():

    farmer_data_insert = []
    farmer_data_insert_row = {}
    farmer_register_file = get_farmer_register_file()
    farmer_register_sheet = get_farmer_register_sheet(farmer_register_file)
    farmer_register = get_farmer_register_titles(farmer_register_sheet)
    for row in farmer_register_sheet.iter_rows(min_row=4, max_col=10, max_row=4, values_only=True):
        farmer_data_list = row
        for i, farmer_data in enumerate(farmer_data_list):
            farmer_register[list(farmer_register)[i]] = farmer_data

        farmer_data_insert_row.update({'name': farmer_register['NOME DO PRODUTOR']})
        farmer_data_insert_row.update({'fantasy_name': farmer_register['NOME DO PRODUTOR']})
        farmer_data_insert_row.update({'cnpo_register': farmer_register['CNPF/CNPJ/NIF']})

        # Entity Type
        entity_type = EntityType()
        for entity_type_itens in entity_type.ENTITY_TYPE:
            if entity_type_itens[0] in farmer_register['TIPO DE ENTIDADE']:
                farmer_data_insert_row.update({'entity_type': EntityType(name=entity_type_itens[0])})
            else:
                # Tratar em caso de erro
                pass

        # Entity
        entity = Entity.objects.all()
        for entity_itens in entity.values():
            if entity_itens['name'] in farmer_register['ENTIDADE']:
                farmer_data_insert_row.update({'entity': entity.get(id=entity_itens['id'])})
            else:
                farmer_data_insert_row.update({'entity': Entity(name=farmer_register['ENTIDADE'])})

        # Address
        geolocator = Nominatim(user_agent="nhanduti")
        location = geolocator.geocode(farmer_register['CIDADE'] + ', ' + farmer_register['UF'] + ', ' + farmer_register['PAIS'])
        address = OrganicFarmerAddress(city=farmer_register['CIDADE'], state=farmer_register['UF'], country=farmer_register['PAIS'], geolocation=Point(location.latitude, location.longitude))
        farmer_data_insert_row.update({'address': address})

        # Activities
        if ';' in farmer_register['ATIVIDADES']:
            activities = set_activity(farmer_register['ATIVIDADES'].split(';'))
        else:
            activities = set_activity(farmer_register['ATIVIDADES'].split(','))

        farmer_data_insert_row.update({'activity': activities})

        # Scopes
        if ';' in farmer_register['ESCOPO']:
            scopes = set_activity(farmer_register['ESCOPO'].split(';'))
        else:
            scopes = set_activity(farmer_register['ESCOPO'].split(','))

        farmer_data_insert_row.update({'scope': scopes})

    farmers.save()
