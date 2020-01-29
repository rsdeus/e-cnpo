from openpyxl import load_workbook
from apps.organic_farmer.models import OrganicFarmerAddress, OrganicFarmerContact, EntityType, Entity, Scope, Activity, \
    OrganicFarmer, CNPOFile
from nhanduti.settings import STATIC_ROOT
from geopy.geocoders import Nominatim
from django.contrib.gis.geos import Point
import re


class GetFarmerRegister:

    farmer_register_file = STATIC_ROOT + '/upload/CNPO_MAPA_31_07_2019.xlsx'

    def get_file_from_mapa(self):
        pass

    def get_farmer_register_ws(self):
        farmer_register_wb = load_workbook(self.farmer_register_file)
        sheet_name = farmer_register_wb.get_sheet_names()[0]
        if sheet_name == 'RELATORIO DE PRODUTOR ORGANICO':
            farmer_register_ws = farmer_register_wb[sheet_name]
            last_row = farmer_register_ws.max_row
            return farmer_register_ws, last_row
        else:
            return False

    def get_farmer_register_headers(self, farmer_register_ws):
        headers = []
        for row in farmer_register_ws.iter_rows(min_row=3, max_col=11, max_row=3, values_only=True):
            headers = list(row)
        return headers

    def set_entity_type(self, farmer_register_entity_type):
        entity_type = EntityType.objects.all()
        entity_type_obj, created = entity_type.get_or_create(name=farmer_register_entity_type)

        return entity_type_obj, created

    def set_entity(self, farmer_register_entity):
        entity = Entity.objects.all()
        entity_obj, created = entity.get_or_create(name=farmer_register_entity)

        return entity_obj, created

    def set_activity(self, farmer_register_activities):
        return_activities_farmer = []
        buscar_activities = []
        criar_cadastro_atividade = []
        for atividade in farmer_register_activities:
            atividade_filter = Activity.objects.filter(name__icontains=atividade.strip().upper())
            if atividade_filter:
                activity_in_db = Activity.objects.get(pk=atividade_filter.values_list()[0][0])
                return_activities_farmer.append(activity_in_db)
            else:
                buscar_activities.append(atividade.strip().upper())
                criar_cadastro_atividade.append(Activity(name=atividade.strip().upper()))
        if criar_cadastro_atividade:
            Activity.objects.bulk_create(criar_cadastro_atividade)
        if buscar_activities:
            return_activities_farmer = self.set_activity(buscar_activities)
            return(return_activities_farmer)
        else:
            return(return_activities_farmer)


    def set_scope(self, farmer_register_scopes):
        return_scope_farmer = []
        buscar_scope = []
        criar_cadastro_scope = []
        escopo_ = ''
        for escopo in farmer_register_scopes:
            if escopo.strip().upper() in ('POV' or 'P.O.V' or 'processador de alimentos'):
                escopo_ = 'PROCESSAMENTO DE PRODUTOS DE ORIGEM VEGETAL'
            elif escopo.strip().upper() in ('POA' or 'P.O.A'):
                escopo_ = 'PROCESSAMENTO DE PRODUTOS DE ORIGEM ANIMAL'
            elif escopo.strip().upper() in ('PPV' or 'P.P.V' or 'produção vegetal'):
                escopo_ = 'PRODUÇÃO PRIMÁRIA VEGETAL'
            elif escopo.strip().upper() in ('PPA' or 'P.P.A'):
                escopo_ = 'PRODUÇÃO PRIMÁRIA ANIMAL'
            elif escopo.strip().upper() in ('ESO' or 'E.S.O' or 'EXT'):
                escopo_ = 'EXTRATIVISMO SUSTENTÁVEL ORGÂNICO'
            elif escopo.strip().upper() in ('PIA' or 'P.I.A'):
                escopo_ = 'PROCESSAMENTO DE INSUMOS AGRÍCOLAS'
            else:
                escopo_ = escopo.strip().upper()

            scope_filter = Scope.objects.filter(name__iexact=escopo_)

            if scope_filter:
                scope_in_db = Scope.objects.get(pk=scope_filter.values_list()[0][0])
                return_scope_farmer.append(scope_in_db)
            else:
                buscar_scope.append(escopo.strip().upper())
                criar_cadastro_scope.append(Scope(name=escopo_.strip().upper()))
        if criar_cadastro_scope:
            try:
                Scope.objects.bulk_create(criar_cadastro_scope)
            except:
                pass
        if buscar_scope:
            return_scope_farmer = self.set_scope(buscar_scope)
            return(return_scope_farmer)
        else:
            return(return_scope_farmer)

    def set_address(self, farmer_register_address):
        address_filter = OrganicFarmerAddress.objects\
            .filter(country__iexact=farmer_register_address['PAIS'])\
            .filter(state__iexact=farmer_register_address['UF'])\
            .filter(city__iexact=farmer_register_address['CIDADE'])

        if address_filter:
            address_obj = OrganicFarmerAddress.objects.get(pk=address_filter.values_list()[0][0])
            created = True
        
        else:         
            geolocator = Nominatim(user_agent="nhanduti")
            address_location = ''
            if farmer_register_address['CIDADE']:
                address_location = farmer_register_address['CIDADE'].strip().upper()
                                       
            if farmer_register_address['UF']:
                address_location += ', ' + farmer_register_address['UF'].strip().upper()

            if farmer_register_address['PAIS']:
                address_location += ', ' + farmer_register_address['PAIS'].strip().upper()

            try:
                location = geolocator.geocode(address_location)

            except:
                location = None

            if location is not None:
                geolocation = Point(location.longitude, location.latitude)
            
            else:
                geolocation = Point(0.0, 0.0)

            address_obj = OrganicFarmerAddress(city=farmer_register_address['CIDADE'],
                                            state=farmer_register_address['UF'],
                                            country=farmer_register_address['PAIS'],
                                            geolocation=geolocation
                                            )

            try:
                address_obj.save()
                created = True

            except:
                created = False

        print(address_obj, created)
        return address_obj, created

    def set_contacts(self, farmer_register_contact):
        pass

    def create_farmer_register(self):

        farmer_data_pre_insert = {}
        farmer_data_insert_row = {}
        farmer_register_ws, last_row = self.get_farmer_register_ws()
        farmer_register_hearders = self.get_farmer_register_headers(farmer_register_ws)
        for row in farmer_register_ws.iter_rows(min_row=4, max_col=10, max_row=1000, values_only=True):
            for i, farmer_data_list in enumerate(row):
                farmer_data_pre_insert.update({farmer_register_hearders[i]: str(farmer_data_list or '').strip().upper()})

            # Name
            farmer_data_insert_row.update({'name': farmer_data_pre_insert['NOME DO PRODUTOR']})


            # CNPO Register
            farmer_data_insert_row.update({'cnpo_register': farmer_data_pre_insert['CNPF/CNPJ/NIF']})

            # Entity Type
            entity_type_obj, created = self.set_entity_type(farmer_data_pre_insert['TIPO DE ENTIDADE'].strip().upper())
            farmer_data_insert_row.update({'entity_type': entity_type_obj})

            # Entity
            entity_obj, created = self.set_entity(farmer_data_pre_insert['ENTIDADE'].strip().upper())
            farmer_data_insert_row.update({'entity': entity_obj})

            # Activities
            if farmer_data_pre_insert['ATIVIDADES']:
                activities_list = []
                for virgula in farmer_data_pre_insert['ATIVIDADES'].split(','):
                    for ponto_virgula in virgula.split(';'):
                        for e in ponto_virgula.split(' e '):
                            for num in re.split(r'[0-9]+', e):
                                activities_list.append(num.strip())
                
                if '' in activities_list:
                    activities_list.remove('')

                activities = self.set_activity(activities_list)

                farmer_data_insert_row.update({'activities': activities})
                print(farmer_data_insert_row['activities'])

            # Scopes
            if farmer_data_pre_insert['ESCOPO']:
                scopes_list = []
                for virgula in farmer_data_pre_insert['ESCOPO'].split(','):
                    for ponto_virgula in virgula.split(';'):
                        for e in ponto_virgula.split(' e '):
                            for num in re.split(r'[0-9]+', e):
                                scopes_list.append(num.strip())
                                print(scopes_list)
                
                if '' in scopes_list:
                    scopes_list.remove('')
                    
                scopes = self.set_scope(scopes_list)

                farmer_data_insert_row.update({'scopes': scopes})

            # Address
            address, created = self.set_address(farmer_data_pre_insert)
            if address:
                farmer_data_insert_row.update({'address': address})
            else:
                pass

            print(farmer_data_insert_row)

            organic_farmer = OrganicFarmer.objects.filter(cnpo_register__icontains=str(farmer_data_insert_row['cnpo_register'] or ''))
            if organic_farmer:
                pass # TODO update no registro
            else:
                update_or_create_organic_farmer = OrganicFarmer.objects.create(name=farmer_data_insert_row['name'],
                                                                cnpo_register=farmer_data_insert_row['cnpo_register'],
                                                                entity=farmer_data_insert_row['entity'],
                                                                address=farmer_data_insert_row['address']
                                                                )
                if 'scopes' in farmer_data_insert_row.keys():
                    for scope in tuple(farmer_data_insert_row['scopes']):
                        update_or_create_organic_farmer.scope.add(scope)

                if 'activities' in farmer_data_insert_row.keys():
                    for activity in tuple(farmer_data_insert_row['activities']):
                        update_or_create_organic_farmer.activities.add(activity)

            # Contacts