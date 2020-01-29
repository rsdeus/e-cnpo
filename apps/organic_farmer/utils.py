from openpyxl import load_workbook


def get_file_from_mapa():
    pass

def load_dataset_file_from_file():

    wb = load_workbook(filename='/home/renato/Projetos/src/nhanduti/staticfiles/upload/copy3_of_CNPO_MAPA_30_11_2019.xlsx', read_only=True)
    sheet_name = wb.get_sheet_names()[0]
    if sheet_name == 'RELATORIO DE PRODUTOR ORGANICO':
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3, max_col=11, max_row=3, values_only=True):
            headers = list(row)
        last_row = ws.max_row
        for row in ws.iter_rows(min_row=3, max_col=11, max_row=last_row, values_only=True):
